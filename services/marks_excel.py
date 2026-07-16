"""
Excel import/export for teacher mark entry.

Template layout:
  Sheet "Info"  — metadata (subject, class, exam, term, year)
  Sheet "Marks" — adm_number | full_name | score
"""

from __future__ import annotations

import io
import re
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill('solid', fgColor='1A6E3C')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
META_LABEL_FONT = Font(bold=True, size=11, color='1A6E3C')
THIN = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)


def build_marks_template(
    *,
    subject_id: int,
    subject_name: str,
    subject_code: str,
    class_name: str,
    exam_type: str,
    term: str,
    year: int,
    students: list,
    existing_marks: Optional[dict] = None,
) -> bytes:
    """
    Build an .xlsx template prefilled with class students.

    students: iterable of objects with adm_number, full_name, id
    existing_marks: optional {student_id: score}
    """
    existing_marks = existing_marks or {}
    wb = Workbook()

    # ── Info sheet (do not delete — used on import) ─────────────────────────
    info = wb.active
    info.title = 'Info'
    info['A1'] = 'Elimu Marks Template'
    info['A1'].font = Font(bold=True, size=14, color='1A6E3C')
    info['A2'] = 'Fill scores on the Marks sheet (0–100). Do not change Info values.'
    info['A2'].font = Font(italic=True, size=10, color='6B7280')

    meta = [
        ('subject_id', subject_id),
        ('subject_name', subject_name),
        ('subject_code', subject_code),
        ('class_name', class_name),
        ('exam_type', exam_type),
        ('term', term),
        ('year', year),
    ]
    info['A4'] = 'Field'
    info['B4'] = 'Value'
    info['A4'].font = HEADER_FONT
    info['B4'].font = HEADER_FONT
    info['A4'].fill = HEADER_FILL
    info['B4'].fill = HEADER_FILL
    for i, (key, val) in enumerate(meta, start=5):
        info[f'A{i}'] = key
        info[f'B{i}'] = val
        info[f'A{i}'].font = META_LABEL_FONT
    info.column_dimensions['A'].width = 18
    info.column_dimensions['B'].width = 28

    # ── Marks sheet ─────────────────────────────────────────────────────────
    marks = wb.create_sheet('Marks')
    headers = ['adm_number', 'full_name', 'score']
    for col, h in enumerate(headers, start=1):
        cell = marks.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN

    for row_i, stud in enumerate(students, start=2):
        marks.cell(row_i, 1, stud.adm_number or '').border = THIN
        marks.cell(row_i, 2, stud.full_name or '').border = THIN
        score = existing_marks.get(stud.id)
        score_cell = marks.cell(row_i, 3, score if score is not None else None)
        score_cell.border = THIN
        score_cell.alignment = Alignment(horizontal='center')

    marks.column_dimensions['A'].width = 16
    marks.column_dimensions['B'].width = 28
    marks.column_dimensions['C'].width = 12

    # Score validation 0–100
    dv = DataValidation(
        type='decimal', operator='between', formula1='0', formula2='100',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid score', error='Score must be between 0 and 100.',
    )
    dv.add(f'C2:C{max(2, len(students) + 1)}')
    marks.add_data_validation(dv)

    marks.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_marks_workbook(file_storage) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """
    Parse uploaded workbook.

    Returns (meta, rows) where rows are {adm_number, full_name, score}.
    Raises ValueError on bad format.
    """
    try:
        wb = load_workbook(file_storage, data_only=True)
    except Exception as ex:
        raise ValueError(f'Could not read Excel file: {ex}') from ex

    if 'Info' not in wb.sheetnames or 'Marks' not in wb.sheetnames:
        raise ValueError('Workbook must contain "Info" and "Marks" sheets. Download the template first.')

    info = wb['Info']
    meta: dict[str, Any] = {}
    for row in info.iter_rows(min_row=5, max_row=20, max_col=2, values_only=True):
        key, val = row[0], row[1]
        if not key:
            continue
        meta[str(key).strip()] = val

    required = ('subject_id', 'class_name', 'exam_type', 'term', 'year')
    missing = [k for k in required if meta.get(k) in (None, '')]
    if missing:
        raise ValueError(f'Missing Info fields: {", ".join(missing)}')

    try:
        meta['subject_id'] = int(meta['subject_id'])
        meta['year'] = int(meta['year'])
    except (TypeError, ValueError) as ex:
        raise ValueError('subject_id and year on Info sheet must be numbers.') from ex

    meta['class_name'] = str(meta['class_name']).strip()
    meta['exam_type'] = str(meta['exam_type']).strip()
    meta['term'] = str(meta['term']).strip()

    marks = wb['Marks']
    headers = [str(c.value).strip().lower() if c.value is not None else '' for c in marks[1]]
    try:
        adm_i = headers.index('adm_number')
        score_i = headers.index('score')
    except ValueError as ex:
        raise ValueError('Marks sheet must have columns: adm_number, full_name, score') from ex
    name_i = headers.index('full_name') if 'full_name' in headers else None

    rows: list[dict[str, Any]] = []
    for row in marks.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        adm = row[adm_i]
        if adm is None or str(adm).strip() == '':
            continue
        adm = str(adm).strip()
        raw_score = row[score_i]
        if raw_score is None or str(raw_score).strip() == '':
            continue  # blank = skip
        try:
            if isinstance(raw_score, str):
                raw_score = raw_score.strip().replace(',', '.')
            score = float(raw_score)
        except (TypeError, ValueError):
            raise ValueError(f'Invalid score for admission number {adm}: {raw_score!r}')
        if score < 0 or score > 100:
            raise ValueError(f'Score out of range for {adm}: {score} (must be 0–100)')
        rows.append({
            'adm_number': adm,
            'full_name': str(row[name_i]).strip() if name_i is not None and row[name_i] else '',
            'score': score,
        })

    if not rows:
        raise ValueError('No scores found. Enter at least one score in the Score column.')

    return meta, rows


def safe_filename(*parts: str) -> str:
    """Build a safe download filename fragment."""
    joined = '_'.join(str(p) for p in parts if p is not None)
    cleaned = re.sub(r'[^A-Za-z0-9._-]+', '_', joined).strip('_')
    return cleaned[:80] or 'marks'
